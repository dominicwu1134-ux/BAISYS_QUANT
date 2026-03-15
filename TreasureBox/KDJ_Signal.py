
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine
import warnings

# ========================
# 1. 忽略 pandas 的 DBAPI2 警告
# ========================
warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy connectable.*")

# ========================
# 2. 数据库连接配置（请替换为你的实际值）
# ========================
DB_CONFIG = {
    'host': '127.0.0.1',
    'port': '5432',
    'database': 'Corenews',
    'user': 'postgres',
    'password': ' '
}

DB_URI = f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"

# ========================
# 3. 计算一个月前的日期（关键：改为30天，信号更精准）
# ========================
one_month_ago = (datetime.today() - timedelta(days=30)).strftime('%Y-%m-%d')
print(f"📅 统计周期：仅使用最近30天（{one_month_ago} 至 今天）的数据，确保信号时效性")

# ========================
# 4. 第一步：从 app_stock_strategy_report 获取有 kdj_signal 的股票（含 stock_code 和 stock_name）
# ========================
print("🔍 正在从 app_stock_strategy_report 提取有 kdj_signal 的股票及其名称...")
engine = create_engine(DB_URI)

query_signal_stocks = f"""
    SELECT DISTINCT
        stock_code,
        stock_name
    FROM app_stock_strategy_report
    WHERE kdj_signal IS NOT NULL
      AND archive_date >= '{one_month_ago}'
    ORDER BY stock_code;
"""

df_signal_stocks = pd.read_sql(query_signal_stocks, engine)
engine.dispose()

if df_signal_stocks.empty:
    print("⚠️ 未找到近30天内有非空 kdj_signal 的股票。")
    exit()


# ✅ 定义前缀函数（Python 层，仅用于数据处理）
def add_exchange_prefix(stock_code):
    stock_str = str(stock_code).strip()
    if stock_str.startswith('6'):
        return f"sh{stock_str}"
    elif stock_str.startswith('0'):
        return f"sz{stock_str}"
    elif stock_str.startswith('8'):
        return f"bj{stock_str}"
    else:
        return f"other{stock_str}"


# 构建股票映射：stock_code -> (stock_name, prefixed_symbol)
stock_info_map = {}
for _, row in df_signal_stocks.iterrows():
    code = row['stock_code']
    name = row['stock_name'] if row['stock_name'] else "未知名称"
    prefixed_symbol = add_exchange_prefix(code)
    stock_info_map[code] = {
        'name': name,
        'symbol': prefixed_symbol
    }

# 获取所有有信号的股票代码
signal_stock_codes = list(stock_info_map.keys())
print(f"✅ 从策略报告中找到 {len(signal_stock_codes)} 只有信号的股票（近30天）")

# ========================
# 5. 第二步：从 stock_daily_kline 获取这些股票的 close 数据，筛选出“有实际价格”的股票
# ========================
print("📥 正在从 stock_daily_kline 查询这些股票的近30天价格数据...")
engine = create_engine(DB_URI)

# 构造符号列表，用于 IN 查询（Python 层构造，安全！）
prefixed_symbols = [stock_info_map[code]['symbol'] for code in signal_stock_codes]
in_clause = "', '".join(prefixed_symbols)

query_kline = f"""
    SELECT DISTINCT symbol
    FROM stock_daily_kline
    WHERE symbol IN ('{in_clause}')
      AND trade_date >= '{one_month_ago}'
      AND "close" IS NOT NULL;
"""

df_valid_symbols = pd.read_sql(query_kline, engine)
engine.dispose()

# ✅ 关键过滤：只保留那些在 kline 中有价格数据的股票
valid_symbols_set = set(df_valid_symbols['symbol'].tolist())
print(f"✅ 在 stock_daily_kline 中找到 {len(valid_symbols_set)} 只有实际价格数据的股票（近30天）")

# 构建“有效股票池”：必须同时有 KDJ 信号 + 价格数据
effective_stock_codes = []
for code in signal_stock_codes:
    symbol = stock_info_map[code]['symbol']
    if symbol in valid_symbols_set:
        effective_stock_codes.append(code)

print(f"🎯 初步有效股票数（有信号 + 有价格）：{len(effective_stock_codes)}")
print(f"📊 已剔除 {len(signal_stock_codes) - len(effective_stock_codes)} 个无价格数据的无效股票")

if len(effective_stock_codes) == 0:
    print("⚠️ 没有股票同时满足‘有KDJ信号’且‘有价格数据’（近30天）。")
    exit()

# ========================
# 6. 第三步：查询每只股票的“最后一次KDJ信号日期”（仅用 stock_code，不依赖函数）
# ========================
print("🔎 正在查询每只股票的最后一次KDJ信号日期...")
engine = create_engine(DB_URI)

query_last_signal_date = f"""
    SELECT
        stock_code,
        MAX(archive_date) AS last_signal_date
    FROM app_stock_strategy_report
    WHERE kdj_signal IS NOT NULL
      AND archive_date >= '{one_month_ago}'
    GROUP BY stock_code
    ORDER BY stock_code;
"""

df_last_signal_date = pd.read_sql(query_last_signal_date, engine)
engine.dispose()

# 构建映射：stock_code -> last_signal_date
last_signal_date_map = {}
for _, row in df_last_signal_date.iterrows():
    code = row['stock_code']
    last_signal_date_map[code] = row['last_signal_date'].strftime('%Y-%m-%d')

# ========================
# 7. 第四步：查询每只股票“信号日”的 close 价格（用 Python 生成 symbol）
# ========================
print("📈 正在获取每只股票在信号日的收盘价...")
# 构造一个包含所有股票代码和对应 symbol 的列表
signal_dates_list = []
for code in effective_stock_codes:
    symbol = stock_info_map[code]['symbol']
    date_str = last_signal_date_map[code]
    signal_dates_list.append((code, symbol, date_str))

# 构造 SQL 查询：批量查多个 (symbol, trade_date)
if not signal_dates_list:
    print("⚠️ 无有效信号日数据。")
    exit()

# 构建 IN 条件：symbol 和 date 的组合
symbol_date_pairs = []
for code, symbol, date_str in signal_dates_list:
    symbol_date_pairs.append(f"('{symbol}', '{date_str}')")

in_clause_dates = ", ".join(symbol_date_pairs)

query_signal_close = f"""
    SELECT 
        symbol,
        trade_date,
        "close"
    FROM stock_daily_kline
    WHERE (symbol, trade_date) IN ({in_clause_dates})
      AND "close" IS NOT NULL;
"""

df_signal_close = pd.read_sql(query_signal_close, engine)
engine.dispose()

# 构建映射：stock_code -> close_on_signal_date
close_on_signal_map = {}
for _, row in df_signal_close.iterrows():
    symbol = row['symbol']
    date_str = row['trade_date'].strftime('%Y-%m-%d')
    close_val = row['close']
    # 反向找 stock_code
    for code, info in stock_info_map.items():
        if info['symbol'] == symbol:
            close_on_signal_map[code] = close_val
            break

print(f"✅ 成功获取 {len(close_on_signal_map)} 只股票的信号日收盘价")

# ========================
# 8. 第五步：获取每只股票的“最新交易日”收盘价
# ========================
print("📈 正在获取每只股票的最新交易日收盘价...")
latest_date_query = f"""
    WITH latest_trades AS (
        SELECT
            symbol,
            MAX(trade_date) AS latest_date
        FROM stock_daily_kline
        WHERE symbol IN ('{in_clause}')
          AND trade_date >= '{one_month_ago}'
          AND "close" IS NOT NULL
        GROUP BY symbol
    )
    SELECT
        lt.symbol,
        lt.latest_date,
        sdk."close" AS latest_close
    FROM latest_trades lt
    JOIN stock_daily_kline sdk ON lt.symbol = sdk.symbol AND lt.latest_date = sdk.trade_date;
"""

df_latest_close = pd.read_sql(latest_date_query, engine)
engine.dispose()

# 构建映射：symbol -> latest_close
latest_close_map = {}
for _, row in df_latest_close.iterrows():
    symbol = row['symbol']
    latest_close = row['latest_close']
    latest_close_map[symbol] = latest_close

# ========================
# 9. 第六步：过滤“信号后价格未上涨”的股票（核心过滤！）
# ========================
print("⚖️ 正在过滤：仅保留‘信号日后股价上涨’的股票...")
final_effective_stock_codes = []
filtered_out = []

for code in effective_stock_codes:
    symbol = stock_info_map[code]['symbol']

    if code not in close_on_signal_map:
        filtered_out.append(f"{stock_info_map[code]['name']} ({symbol}) - 无信号日价格")
        continue

    if symbol not in latest_close_map:
        filtered_out.append(f"{stock_info_map[code]['name']} ({symbol}) - 无最新价格")
        continue

    signal_close = close_on_signal_map[code]
    latest_close = latest_close_map[symbol]

    # ✅ 核心过滤：最新收盘价 > 信号日收盘价
    if latest_close > signal_close:
        final_effective_stock_codes.append(code)
    else:
        filtered_out.append(
            f"{stock_info_map[code]['name']} ({symbol}) - 信号日 {signal_close}，最新价 {latest_close} ❌ 未上涨")

print(f"✅ 最终通过趋势验证的有效股票：{len(final_effective_stock_codes)} 只")
print(f"❌ 被过滤掉的股票：{len(filtered_out)} 只")
for reason in filtered_out:
    print(f"   - {reason}")

if len(final_effective_stock_codes) == 0:
    print("⚠️ 没有任何股票满足‘KDJ信号后股价上涨’的条件。")
    exit()

# 更新为最终有效股票池
stock_info_map = {code: stock_info_map[code] for code in final_effective_stock_codes}
prefixed_stock_symbols = [stock_info_map[code]['symbol'] for code in final_effective_stock_codes]

print(
    f"✅ 最终有效股票列表：{[f'{stock_info_map[code]['name']} ({stock_info_map[code]['symbol']})' for code in final_effective_stock_codes[:5]]}...")

# ========================
# 新增：计算每只股票的信号日到最新日的涨幅百分比
# ========================
print("📈 正在计算每只股票的 KDJ 信号日到最新日的涨幅百分比...")
gain_percentage_map = {}

for code in final_effective_stock_codes:
    symbol = stock_info_map[code]['symbol']
    signal_close = close_on_signal_map[code]
    latest_close = latest_close_map[symbol]

    # 计算涨幅百分比：(最新价 - 信号价) / 信号价 * 100
    gain_pct = ((latest_close - signal_close) / signal_close) * 100
    gain_percentage_map[code] = round(gain_pct, 2)  # 保留两位小数

print(f"✅ 已计算 {len(gain_percentage_map)} 只股票的涨幅百分比")

# ========================
# 10. 第七步：从 stock_daily_kline 拉取这1个月内所有有效股票的 close 数据
# ========================
print("📥 正在从 stock_daily_kline 拉取近30天的收盘数据（仅限最终有效股票）...")
engine = create_engine(DB_URI)

in_clause = "', '".join(prefixed_stock_symbols)
query_kline = f"""
    SELECT 
        symbol,
        trade_date,
        "close"
    FROM stock_daily_kline
    WHERE symbol IN ('{in_clause}')
      AND trade_date >= '{one_month_ago}'
      AND "close" IS NOT NULL
    ORDER BY symbol, trade_date;
"""

df_kline = pd.read_sql(query_kline, engine)
engine.dispose()

if df_kline.empty:
    print("⚠️ 未找到有效股票的交易数据（近30天）。")
    exit()

# 转换日期格式，统一为 YYYY-MM-DD
df_kline['trade_date'] = pd.to_datetime(df_kline['trade_date']).dt.strftime('%Y-%m-%d')
trade_dates = sorted(df_kline['trade_date'].unique())
print(f"✅ 共获取 {len(trade_dates)} 个交易日，覆盖范围：{trade_dates[0]} 至 {trade_dates[-1]}")

# 构建映射：symbol → {date: close}
close_map = {}
for _, row in df_kline.iterrows():
    symbol = row['symbol']
    date_str = row['trade_date']
    close_map.setdefault(symbol, {})[date_str] = row['close']

# ========================
# 11. 第八步：查询每个股票在哪些日期有 kdj_signal（用于高亮）
# ========================
print("🔎 正在查询各股票的 kdj_signal 信号日（近30天）...")
engine = create_engine(DB_URI)

query_signals = f"""
    SELECT 
        stock_code,
        archive_date::date AS archive_date
    FROM app_stock_strategy_report
    WHERE kdj_signal IS NOT NULL
      AND archive_date >= '{one_month_ago}'
    ORDER BY stock_code, archive_date;
"""

df_signals = pd.read_sql(query_signals, engine)
engine.dispose()

# 构建高亮映射：(symbol, date_str) → True
highlight_map = {}
for _, row in df_signals.iterrows():
    stock_code = row['stock_code']
    date_str = row['archive_date'].strftime('%Y-%m-%d')
    prefixed_symbol = add_exchange_prefix(stock_code)
    # 只有在最终有效股票池中的才高亮
    if prefixed_symbol in prefixed_stock_symbols:
        highlight_map[(prefixed_symbol, date_str)] = True

print(f"✅ 共 {len(highlight_map)} 个 KDJ 信号点可用于高亮（近30天，且通过趋势验证）")

# ========================
# 11.5 新增：查询每个股票的 macd_12269_signal 信号日（仅对已筛选股票做高亮）
# ========================
print("🔎 正在查询各股票的 macd_12269_signal 信号日（仅对已筛选股票高亮）...")

query_macd_signals = f"""
    SELECT 
        stock_code,
        archive_date::date AS archive_date,
        macd_12269_signal
    FROM app_stock_strategy_report
    WHERE macd_12269_signal IS NOT NULL
      AND archive_date >= '{one_month_ago}'
    ORDER BY stock_code, archive_date;
"""

df_macd_signals = pd.read_sql(query_macd_signals, engine)
engine.dispose()

# 构建 MACD 高亮映射：(symbol, date_str) → color_hex
# 注意：只对“已进入最终报表”的股票做高亮，不新增股票
macd_highlight_map = {}

for _, row in df_macd_signals.iterrows():
    stock_code = row['stock_code']
    date_str = row['archive_date'].strftime('%Y-%m-%d')
    signal_value = str(row['macd_12269_signal']).strip()

    prefixed_symbol = add_exchange_prefix(stock_code)

    # ⚠️ 关键：只对已筛选股票做高亮
    if prefixed_symbol not in prefixed_stock_symbols:
        continue

    # 判断 MACD 信号类型（请根据你的数据库实际值调整！）
    if signal_value in ['下金叉', '下叉', '金叉', 'buy', '1', 'BUY', '正金叉']:
        color = "ADD8E6"  # 浅蓝色：零轴下金叉
        macd_highlight_map[(prefixed_symbol, date_str)] = color
    elif signal_value in ['上金叉', '上叉', '死叉', 'sell', '-1', 'SELL', '负金叉']:
        color = "9370DB"  # 紫色：零轴上金叉
        macd_highlight_map[(prefixed_symbol, date_str)] = color
    else:
        # 未知信号，忽略
        continue

print(f"✅ 共 {len(macd_highlight_map)} 个 MACD 信号点可用于高亮（仅限已筛选股票）")

# ========================
# 12. 生成 Excel：横向是交易日，纵向是股票（名称+代码），值是 close，信号日高亮
# ========================
print("🎨 正在生成 Excel 报告（仅含‘有信号 + 有价格 + 信号后上涨’的股票）...")

wb = Workbook()
ws = wb.active
ws.title = "Stock Close & Signal Report"

# ——————————————————————————————————————
# 12.1 插入标题行
# ——————————————————————————————————————
ws.insert_rows(1)
title_cell = ws.cell(row=1, column=1, value="股票收盘价与多因子信号聚焦报告（近30天，仅展示有信号且价格持续上涨的股票）")
title_cell.font = Font(bold=True, size=16, color="2E5488")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(trade_dates) + 2)
ws.row_dimensions[1].height = 35

# ——————————————————————————————————————
# 12.2 插入前置说明（图例 + 筛选逻辑）
# ——————————————————————————————————————
ws.insert_rows(2)
note_cell = ws.cell(row=2, column=1,
                    value="📌 筛选逻辑：仅展示‘有KDJ信号’且‘信号后股价上涨’的股票。\n"
                          "🎨 高亮说明：\n"
                          "🔵 蓝色：MACD 零轴下金叉（买入信号）\n"
                          "🟣 紫色：MACD 零轴上金叉（买入信号）\n"
                          "🔴 红色：KDJ 信号（买入/卖出）\n"
                          "✅ 所有股票均满足：信号日后价格上涨，确保动能有效。")
note_cell.font = Font(bold=True, color="2E5488", size=12)
note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=len(trade_dates) + 2)
ws.row_dimensions[2].height = 80  # 足够高度展示多行图例

# ——————————————————————————————————————
# 12.3 写入表头：第一列是“股票名称 (代码)”，第二列是涨幅%，后续是交易日期
# ——————————————————————————————————————
ws.cell(row=5, column=1, value="Stock Name (Code)")
ws.cell(row=5, column=2, value="Signal to Latest Gain (%)")  # ← 新增涨幅列
for col_idx, date_str in enumerate(trade_dates, 3):  # ← 从第3列开始写日期
    ws.cell(row=5, column=col_idx, value=date_str)

# ——————————————————————————————————————
# 12.4 写入数据：每只有效股票一行，展示涨幅% + 所有交易日的 close
# ——————————————————————————————————————
row_idx = 6
for stock_code in final_effective_stock_codes:
    name = stock_info_map[stock_code]['name']
    symbol = stock_info_map[stock_code]['symbol']
    display_name = f"{name} ({symbol})"

    # 第1列：股票名称
    ws.cell(row=row_idx, column=1, value=display_name)

    # 第2列：涨幅百分比（带正负号，保留两位小数，加%）
    gain_pct = gain_percentage_map[stock_code]
    ws.cell(row=row_idx, column=2, value=f"{gain_pct:+.2f}%")

    # 第3列起：交易日收盘价
    for col_idx, date_str in enumerate(trade_dates, 3):
        close_val = close_map.get(symbol, {}).get(date_str, None)
        cell = ws.cell(row=row_idx, column=col_idx, value=close_val)

        # ✅ 优先应用 MACD 信号颜色
        if (symbol, date_str) in macd_highlight_map:
            color_hex = macd_highlight_map[(symbol, date_str)]
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            cell.fill = fill
        # ✅ 否则，应用 KDJ 信号颜色（红色）
        elif (symbol, date_str) in highlight_map:
            fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.fill = fill
        # ✅ 否则：无高亮

    row_idx += 1

# ——————————————————————————————————————
# 12.5 自动调整列宽
# ——————————————————————————————————————
print("📏 正在自动调整列宽...")
total_cols = len(trade_dates) + 2  # 原来是 +1，现在是 +2（多了涨幅列）

for col_idx in range(1, total_cols + 1):
    max_length = 0
    column = get_column_letter(col_idx)

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell_len = len(str(cell.value))
            if cell_len > max_length:
                max_length = cell_len

    adjusted_width = min(max(max_length + 2, 8), 25)
    ws.column_dimensions[column].width = adjusted_width

# ——————————————————————————————————————
# 12.6 保存文件
# ——————————————————————————————————————
output_file = "stock_close_kdj_macd_signal_report.xlsx"
wb.save(output_file)
print(f"🎉 Excel 文件已生成：{output_file}")
print(f"📊 最终报告包含：{len(final_effective_stock_codes)} 只‘信号+价格+上涨动能’三重确认的股票。")
print(f"⏱️ 报告周期：仅使用最近30天数据，确保信号高时效性与交易价值。")
print(f"✅ 每行都是可交易标的，无噪音、无陷阱、无过期信号。")
